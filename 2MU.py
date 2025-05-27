import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from copy import copy

# 读取excel_utils模块，包含所需的函数
from excel_utils import (
    extract_birthday_and_add_to_column,
    split_info_to_next_row,
    split_column_and_add,
    convert_names_to_pinyin,
    save_grouped_to_sheets
)
def split_sheets_to_individual_files(output_file_path, output_dir):
    """拆分每个工作表成独立的Excel文件，文件名格式为 MU_工作表名称_A3单元格内容，并删除A列"""
    workbook = load_workbook(output_file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # 获取 A3 单元格内容
        a3_content = sheet["A3"].value if sheet["A3"].value else "Empty"

        # 创建文件名
        file_name = f"MU_{sheet_name}_{a3_content}.xlsx"
        file_name = file_name.replace("/", "-")  # 防止非法字符
        file_path = os.path.join(output_dir, file_name)

        # 创建新的工作簿
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name

        # 复制原始工作表的内容到新的工作表，包括值、样式和合并单元格
        copy_sheet(sheet, new_sheet)

        # 取消可能存在的合并单元格
        if new_sheet.merged_cells.ranges:
            merged_cells = list(new_sheet.merged_cells)
            for merged_cell in merged_cells:
                new_sheet.unmerge_cells(str(merged_cell))

        # 删除 A 列
        new_sheet.delete_cols(1)

        # 获取最大列数
        max_col = new_sheet.max_column

        # 重新设置首行格式
        new_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)  # 合并A1到C1
        new_sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)  # 合并D1到G1

        # 检查是否存在H列和I列
        if max_col >= 8:  # H列的索引是8
            if max_col >= 9:  # I列的索引是9
                new_sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)  # 合并H1到I1
            else:
                # 只有H列存在，不需要合并
                pass
        else:
            # H列不存在，跳过合并
            pass

        # 设置首行内容
        new_sheet.cell(row=1, column=1, value="姓名信息(中英文至少填写一项）")
        new_sheet.cell(row=1, column=4, value="证件信息（至少填写一种证件）")

        if max_col >= 8:
            new_sheet.cell(row=1, column=8, value="C0客户必填")
        else:
            pass  # H列不存在，跳过设置

        # 设置首行字体和对齐方式
        for col in range(1, max_col + 1):
            cell = new_sheet.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="宋体", bold=True)

        # 设置列宽
        for col_idx, width in zip(range(1, max_col + 1), [28.75]*3 + [28.5]*4 + [15.5]*2):
            if col_idx <= max_col:
                col_letter = get_column_letter(col_idx)
                new_sheet.column_dimensions[col_letter].width = width

        # 设置行高
        new_sheet.row_dimensions[1].height = 23

        new_workbook.save(file_path)
        print(f"保存独立文件并删除A列：{file_path}")


def copy_sheet(source_sheet, target_sheet):
    """复制工作表，包括单元格的值、样式、合并单元格"""
    # 复制单元格
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    # 复制行高和列宽
    for row_idx, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = row_dim.height
    for col_idx, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_idx].width = col_dim.width
    # 复制合并单元格
    for merged_cell in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell))

def modify_sheets(output_file_path):
    workbook = load_workbook(output_file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 取消所有合并单元格
        if sheet.merged_cells.ranges:
            merged_cells = list(sheet.merged_cells)
            for merged_cell in merged_cells:
                sheet.unmerge_cells(str(merged_cell))

        # 插入新行
        sheet.insert_rows(1)

        # 在新行的B1, E1, I1设置内容
        sheet["B1"] = "姓名信息(中英文至少填写一项）"
        sheet["E1"] = "证件信息（至少填写一种证件）"
        sheet["I1"] = "C0客户必填"

        # 合并单元格
        sheet.merge_cells("B1:D1")
        sheet.merge_cells("E1:H1")
        sheet.merge_cells("I1:J1")

        # 设置字体样式
        for cell_range in ["B1:D1", "E1:H1", "I1:J1"]:
            for row in sheet[cell_range]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="宋体", bold=True)

        # 设置新行的行高
        sheet.row_dimensions[1].height = 23

        # 修改原来的第一行（现在是第二行）为新内容
        headers = [
            "员工姓名（中）", "员工姓名（英/拼音）", "生日", "身份证号码",
            "护照号码", "其他证件类型（下拉选择）", "其他证件号", "所属企业名称", "企业所在地"
        ]
        for col_idx, header in enumerate(headers, start=2):
            cell = sheet.cell(row=2, column=col_idx)  # 修改第二行
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="宋体", bold=True)

        # 设置第二行的行高为34.5
        sheet.row_dimensions[2].height = 34.5

        # 设置列宽
        for col, width in zip(["B", "C", "D"], [28.75, 28.75, 28.75]):
            sheet.column_dimensions[col].width = width
        for col, width in zip(["E", "F", "G", "H"], [28.5, 28.5, 28.5, 28.5]):
            sheet.column_dimensions[col].width = width
        for col, width in zip(["I", "J"], [15.5, 15.5]):
            sheet.column_dimensions[col].width = width

        # 设置E, F, H列单元格格式为纯文本
        for col_letter in ["E", "F", "H"]:
            for row in sheet.iter_rows(min_col=column_index_from_string(col_letter),
                                       max_col=column_index_from_string(col_letter),
                                       min_row=3):
                for cell in row:
                    cell.number_format = FORMAT_TEXT

        # 设置颜色填充
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

        for col_letter in ["B", "C", "E", "F", "G", "H"]:
            sheet[f"{col_letter}2"].fill = red_fill  # B2, C2, E2-H2填充红色
        for col_letter in ["D", "I", "J"]:
            sheet[f"{col_letter}2"].fill = yellow_fill  # D2, I2, J2填充黄色

        # 删除G3:L3及以下单元格内容
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=7, max_col=12):  # G列到L列
            for cell in row:
                cell.value = None

        # 筛选替换逻辑
        max_row = sheet.max_row  # 由于后续可能删除行，需要提前获取最大行数
        for row_index in range(3, max_row + 1):
            n_value = sheet[f"N{row_index}"].value
            if n_value == "身份证":  # 检测N列内容是否为"身份证"
                id_value = sheet[f"M{row_index}"].value  # 获取对应M列单元格的值
                sheet[f"E{row_index}"].value = id_value  # 将M列值复制到E列
                sheet[f"C{row_index}"].value = None  # 清空C列单元格内容
                sheet[f"D{row_index}"].value = None  # 清空D列单元格内容
                sheet[f"F{row_index}"].value = None  # 清空F列单元格内容

            elif n_value in ["普通护照", "公务护照"]:  # 检测N列内容是否为护照
                c_value = sheet[f"C{row_index}"].value or ""
                d_value = sheet[f"D{row_index}"].value or ""
                # 合并C列和D列内容，转大写并去除空格
                combined_value = (str(c_value).replace(" ", "") + "/" + str(d_value).replace(" ", "")).upper()# 合并C列和D列内容，中间添加"/"，转大写并去除空格
                sheet[f"C{row_index}"].value = combined_value  # 设置合并后的值到C列
                sheet[f"D{row_index}"].value = sheet[f"E{row_index}"].value  # 将E列内容复制到D列
                sheet[f"F{row_index}"].value = sheet[f"M{row_index}"].value  # 将M列内容复制到F列
                sheet[f"E{row_index}"].value = None  # 删除E列单元格内容

            elif n_value:  # 如果N列内容既不是身份证也不是护照类型
                sheet[f"G{row_index}"].value = n_value  # 将N列内容复制到G列
                sheet[f"H{row_index}"].value = sheet[f"M{row_index}"].value  # 将M列内容复制到H列
                # 合并C列和D列内容，转大写去空格
                c_value = sheet[f"C{row_index}"].value or ""
                d_value = sheet[f"D{row_index}"].value or ""
                combined_value = (str(c_value).replace(" ", "") + "/" + str(d_value).replace(" ", "")).upper()# 合并C列和D列内容，中间添加"/"，转大写并去除空格
                sheet[f"C{row_index}"].value = combined_value
                sheet[f"D{row_index}"].value = sheet[f"E{row_index}"].value  # 将E列内容复制到D列
                # 删除F列内容
                sheet[f"F{row_index}"].value = None
                sheet[f"E{row_index}"].value = None  # 删除E列单元格内容

        # 检查B列相邻单元格内容，当N列内容没有“身份证”时合并相邻行
        row = 3
        while row < sheet.max_row:
            b_value = sheet[f"B{row}"].value
            n_value = sheet[f"N{row}"].value
            b_next = sheet[f"B{row + 1}"].value
            n_next = sheet[f"N{row + 1}"].value

            if b_value == b_next and "身份证" not in [n_value, n_next]:
                for col in range(2, sheet.max_column + 1):  # 遍历列
                    current_cell = sheet.cell(row=row, column=col)
                    next_cell = sheet.cell(row=row + 1, column=col)

                    if current_cell.value is None and next_cell.value is not None:
                        current_cell.value = next_cell.value  # 将下一行内容合并到当前行
                sheet.delete_rows(row + 1)  # 删除下一行
                continue  # 继续检查当前行
            row += 1

        # 检查B列相邻单元格内容，当相邻行N列内容有“身份证”时保留身份证行
        rows_to_delete = []
        row = 3
        while row < sheet.max_row:
            b_value = sheet[f"B{row}"].value
            n_value = sheet[f"N{row}"].value
            b_next = sheet[f"B{row + 1}"].value
            n_next = sheet[f"N{row + 1}"].value

            if b_value == b_next:  # 如果B列内容相邻一致
                if "身份证" in [n_value, n_next]:  # 如果N列中有身份证
                    if n_value == "身份证":
                        rows_to_delete.append(row + 1)  # 删除下一行
                    elif n_next == "身份证":
                        rows_to_delete.append(row)  # 删除当前行
                    row += 1  # 跳过下一行
            row += 1

        # 删除收集到的行
        for row in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row)

        # 删除B列单元格内容，当对应C列单元格有内容时
        for row in range(3, sheet.max_row + 1):
            if sheet[f"C{row}"].value:  # 如果C列有内容
                sheet[f"B{row}"].value = None  # 删除对应B列内容

        # 删除J-M列
        sheet.delete_cols(11, 4)  # 从J列开始删除4列

    # 保存修改
    workbook.save(output_file_path)

def main():
    input_file = r"请替换为你实际的路径\RawData\MUwhitelist_updated.xlsx"
    output_dir = r"请替换为你实际的路径\output"
    output_file_name = "MU协议号拆分.xlsx"

    if not os.path.exists(input_file):
        print(f"输入文件不存在：{input_file}")
        return

    df = pd.read_excel(input_file)

    if '公司名称' not in df.columns or df['公司名称'].isnull().any():
        print("警告：公司名称列缺失或存在空值，请检查数据！")
        return

    # 数据处理
    df = extract_birthday_and_add_to_column(df)
    df = split_info_to_next_row(df)
    df = split_column_and_add(df)
    df = convert_names_to_pinyin(df)

    # 保存到单一文件，分组数据存入独立工作表
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_file_path = os.path.join(output_dir, output_file_name)
    save_grouped_to_sheets(
        df,
        save_path=output_dir,
        file_name=output_file_name,
        company_name_col='公司名称',
        agreement_col='协议号'
    )

    # 修改所有工作表
    modify_sheets(output_file_path)

    # 拆分工作表成独立文件并删除A列
    split_sheets_to_individual_files(output_file_path, output_dir)

    print("处理完成！")

if __name__ == "__main__":
    main()
