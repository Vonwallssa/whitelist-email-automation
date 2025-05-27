import pandas as pd
from datetime import datetime
from pypinyin import lazy_pinyin
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_TEXT
import os
import re

# 获取拼音
def get_char_pinyin(char):
    return ''.join(lazy_pinyin(char))

# 提取身份证生日
def extract_birthday_and_add_to_column(df, id_col='证件信息', birthday_col='员工生日'):
    def extract_birthday_from_id(value):
        if isinstance(value, str) and '身份证' in value:
            parts = value.split('|')
            return parts[1][6:14] if len(parts) > 1 and len(parts[1]) == 18 else None
        return None

    def format_existing_birthday(value):
        try:
            return datetime.strptime(value, '%Y-%m-%d').strftime('%Y%m%d') if isinstance(value, str) and '-' in value else value
        except ValueError:
            return value

    df[birthday_col] = df[id_col].apply(extract_birthday_from_id).fillna(df[birthday_col].apply(format_existing_birthday))
    return df

# 拆分信息列
def split_info_to_next_row(df, col='证件信息'):
    if col in df.columns:
        expanded_rows = df[col].str.split(',', expand=True).stack().reset_index(level=1, drop=True).to_frame(col)
        df = df.drop(columns=[col]).join(expanded_rows).reset_index(drop=True)
    return df

# 拆分列并新增列
def split_column_and_add(df, col='证件信息', new_col='证件类型'):
    if col in df.columns:
        split_data = df[col].str.split('|', expand=True)
        df[new_col] = split_data[0]
        df[col] = split_data[1]
    return df

# 将姓名拆分为拼音
def convert_names_to_pinyin(df, name_col='姓名', surname_col='姓', givenname_col='名'):
    def split_name(name):
        if isinstance(name, str) and all('\u4e00' <= char <= '\u9fff' for char in name):
            surname = get_char_pinyin(name[0]).upper()
            givenname = ''.join(get_char_pinyin(char).upper() for char in name[1:])
            return pd.Series([surname, givenname])
        return pd.Series([None, None])

    if name_col in df.columns:
        df[[surname_col, givenname_col]] = df[name_col].apply(split_name)
    return df

# 清理字符串内容
def clean_string(s):
    if not s:
        return "未知公司"
    invalid_chars = r'[:\/<>|"?*\t]'
    return re.sub(invalid_chars, '', str(s))[:31]

# 保存到独立工作表
def save_grouped_to_sheets(df, save_path, file_name, company_name_col='公司名称', agreement_col='协议号', add_and_merge_header=None, set_header_titles_and_format=None):
    grouped = df.groupby(agreement_col)  # 按协议号分组

    # 创建新的工作簿
    wb = Workbook()
    first_sheet = True

    for agreement_value, group in grouped:
        sheet_name = clean_string(str(agreement_value))[:31]  # Excel工作表名最大长度为31字符

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row in dataframe_to_rows(group, index=False, header=True):
            ws.append(row)

        if add_and_merge_header:
            add_and_merge_header(ws)
        if set_header_titles_and_format:
            set_header_titles_and_format(ws)

    file_path = os.path.join(save_path, file_name)
    wb.save(file_path)
    print(f"保存文件：{file_path}")